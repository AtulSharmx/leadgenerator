from fastapi import FastAPI, HTTPException, Response, Request
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List
from motor.motor_asyncio import AsyncIOMotorClient
from datetime import datetime, timezone, timedelta
from dotenv import load_dotenv
import os
import uuid
import httpx
import io
import csv

# Load environment variables
load_dotenv()

app = FastAPI(title="Lead Generator API")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# MongoDB
MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "leadradar")
FOURSQUARE_API_KEY = os.environ.get("FOURSQUARE_API_KEY")

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

# Search cache
search_cache = {}

# Foursquare API function
async def search_foursquare(city: str, niche: str) -> List[dict]:
    """Search businesses using Foursquare Places API"""
    if not FOURSQUARE_API_KEY:
        print("No Foursquare API key configured - using mock data")
        return []
    
    try:
        async with httpx.AsyncClient() as client:
            # Search for places
            search_url = "https://api.foursquare.com/v3/places/search"
            headers = {
                "Authorization": FOURSQUARE_API_KEY,
                "Accept": "application/json"
            }
            params = {
                "query": niche,
                "near": city,
                "limit": 50,
                "fields": "fsq_id,name,location,tel,website,rating,stats,email,hours"
            }
            
            response = await client.get(search_url, headers=headers, params=params)
            
            if response.status_code == 401:
                print(f"Foursquare API: Invalid API key - using mock data")
                return []
            
            if response.status_code != 200:
                print(f"Foursquare API error: {response.status_code} - {response.text}")
                return []
            
            data = response.json()
            places = data.get("results", [])
            
            businesses = []
            for place in places:
                location = place.get("location", {})
                stats = place.get("stats", {})
                
                # Build address
                address_parts = []
                if location.get("address"):
                    address_parts.append(location["address"])
                if location.get("locality"):
                    address_parts.append(location["locality"])
                if location.get("region"):
                    address_parts.append(location["region"])
                if location.get("country"):
                    address_parts.append(location["country"])
                
                full_address = ", ".join(address_parts) if address_parts else city
                
                website = place.get("website")
                
                businesses.append({
                    "id": str(uuid.uuid4()),
                    "name": place.get("name", "Unknown"),
                    "phone": place.get("tel", "N/A"),
                    "address": full_address,
                    "email": place.get("email"),
                    "website": website,
                    "hasWebsite": bool(website),
                    "rating": place.get("rating", 0) / 2,  # Foursquare uses 0-10, convert to 0-5
                    "totalReviews": stats.get("total_ratings", 0),
                    "category": niche,
                    "placeId": place.get("fsq_id", "")
                })
            
            print(f"Foursquare API: Found {len(businesses)} businesses for '{niche}' in '{city}'")
            return businesses
            
    except Exception as e:
        print(f"Foursquare API error: {e}")
        return []

# Models
class SearchRequest(BaseModel):
    city: str
    niche: str

class SessionRequest(BaseModel):
    sessionId: str

class ExportRequest(BaseModel):
    businesses: List[dict]
    city: Optional[str] = None
    niche: Optional[str] = None

class Business(BaseModel):
    id: str
    name: str
    phone: str
    address: str
    email: Optional[str]
    website: Optional[str]
    hasWebsite: bool
    rating: float
    totalReviews: int
    category: str
    placeId: str

# Mock data generator
def generate_mock_businesses(city: str, niche: str, count: int = 50) -> List[dict]:
    import random
    
    name_templates = {
        'Gym': ['FitZone', 'PowerHouse Gym', 'Iron Paradise', 'Elite Fitness', 'CrossFit', 'Muscle Factory', 'Body Works', 'The Gym', 'Fitness First', 'Active Life'],
        'Salon': ['Glamour Studio', 'Style Hub', 'Beauty Bliss', 'Hair Art', 'The Salon', 'Elegance', 'Chic & Style', 'Glow Beauty', 'Perfect Look', 'Radiance'],
        'Restaurant': ['The Kitchen', 'Spice Garden', 'Urban Bites', 'Taste of India', 'Golden Dragon', 'Pasta Palace', 'Grill House', 'Food Factory', 'Dine Divine', 'Flavor Town'],
        'Real Estate': ['Prime Properties', 'Dream Homes', 'City Realty', 'Property Hub', 'HomeFind', 'Estate Masters', 'Land Mark', 'Brick & Mortar', 'Key Realty', 'Urban Estates'],
        'Hospital': ['City Hospital', 'Care Plus', 'Health First', 'Metro Medical', 'Life Care', 'Wellness Center', 'Medical Hub', 'Apollo Care', 'Fortis Health', 'Max Healthcare'],
        'Hotel': ['Grand Plaza', 'Royal Inn', 'Comfort Stay', 'The Residence', 'Park View', 'Luxury Suites', 'City Lodge', 'Heritage Hotel', 'Premier Inn', 'Golden Tulip'],
        'School': ['Cambridge Academy', "St. Mary's", 'Public School', 'International School', 'Learning Hub', 'Knowledge Academy', 'Scholars School', 'Bright Future', 'Elite Academy', 'Modern School'],
        'Clinic': ['Health Clinic', 'Care Clinic', 'Family Health', 'Wellness Clinic', 'Medical Care', 'Quick Care', 'City Clinic', 'Health Point', 'Life Clinic', 'Prime Health']
    }
    
    street_names = ['Main Street', 'Park Avenue', 'Market Road', 'Station Road', 'MG Road', 'Ring Road', 'Highway', 'Gandhi Nagar', 'Civil Lines', 'Sector']
    phone_prefixes = ['98', '99', '97', '96', '95', '91', '90', '88', '87', '70']
    
    names = name_templates.get(niche, name_templates['Restaurant'])
    businesses = []
    
    for i in range(count):
        base_name = random.choice(names)
        has_website = random.random() < 0.6
        has_email = random.random() < 0.4
        rating = round(3.5 + random.random() * 1.5, 1)
        reviews = random.randint(10, 500)
        phone = f"+91 {random.choice(phone_prefixes)}{random.randint(10000000, 99999999)}"
        street = random.choice(street_names)
        number = random.randint(1, 500)
        clean_name = ''.join(c.lower() for c in base_name if c.isalpha())
        
        businesses.append({
            "id": str(uuid.uuid4()),
            "name": f"{base_name} {city.split()[0]}{f' {i+1}' if i > 0 else ''}",
            "phone": phone,
            "address": f"{number}, {street}, {city}",
            "email": f"contact@{clean_name}.com" if has_email else None,
            "website": f"https://www.{clean_name}.com" if has_website else None,
            "hasWebsite": has_website,
            "rating": rating,
            "totalReviews": reviews,
            "category": niche,
            "placeId": f"mock_{uuid.uuid4()}"
        })
    
    return businesses

# Routes
@app.get("/api/health")
async def health_check():
    return {"status": "ok", "timestamp": datetime.now(timezone.utc).isoformat()}

@app.post("/api/search")
async def search_businesses(request: SearchRequest):
    city = request.city
    niche = request.niche
    
    if not city or not niche:
        raise HTTPException(status_code=400, detail="City and niche are required")
    
    cache_key = f"{city.lower()}_{niche.lower()}"
    
    # Check cache
    if cache_key in search_cache:
        return search_cache[cache_key]
    
    # Try Foursquare API first, fall back to mock data
    businesses = await search_foursquare(city, niche)
    
    if not businesses:
        # Fallback to mock data if API fails or no key
        businesses = generate_mock_businesses(city, niche)
    
    result = {
        "businesses": businesses,
        "total": len(businesses),
        "hasWebsite": len([b for b in businesses if b["hasWebsite"]]),
        "noWebsite": len([b for b in businesses if not b["hasWebsite"]]),
        "city": city,
        "niche": niche,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    
    # Cache results
    search_cache[cache_key] = result
    
    # Store search in DB
    await db.searches.insert_one({
        "search_id": str(uuid.uuid4()),
        "city": city,
        "niche": niche,
        "resultCount": len(businesses),
        "createdAt": datetime.now(timezone.utc)
    })
    
    return result

@app.post("/api/export/csv")
async def export_csv(request: ExportRequest):
    businesses = request.businesses
    
    if not businesses:
        raise HTTPException(status_code=400, detail="No businesses to export")
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(['Name', 'Phone', 'Email', 'Address', 'Website', 'Rating', 'Reviews', 'Category'])
    
    # Data
    for b in businesses:
        writer.writerow([
            b.get('name', ''),
            b.get('phone', ''),
            b.get('email', 'N/A'),
            b.get('address', ''),
            b.get('website', 'No Website'),
            b.get('rating', ''),
            b.get('totalReviews', ''),
            b.get('category', '')
        ])
    
    csv_content = output.getvalue()
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=leads.csv"}
    )

@app.post("/api/export/excel")
async def export_excel(request: ExportRequest):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export not available")
    
    businesses = request.businesses
    
    if not businesses:
        raise HTTPException(status_code=400, detail="No businesses to export")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    
    # Header
    headers = ['Business Name', 'Phone', 'Email', 'Address', 'Website', 'Rating', 'Reviews', 'Category']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF6B2B", end_color="FF6B2B", fill_type="solid")
    
    # Data
    for row, b in enumerate(businesses, 2):
        ws.cell(row=row, column=1, value=b.get('name', ''))
        ws.cell(row=row, column=2, value=b.get('phone', ''))
        ws.cell(row=row, column=3, value=b.get('email') or 'N/A')
        ws.cell(row=row, column=4, value=b.get('address', ''))
        ws.cell(row=row, column=5, value=b.get('website') or 'No Website')
        ws.cell(row=row, column=6, value=b.get('rating', ''))
        ws.cell(row=row, column=7, value=b.get('totalReviews', ''))
        ws.cell(row=row, column=8, value=b.get('category', ''))
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 30
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=leads.xlsx"}
    )

@app.post("/api/export/pdf")
async def export_pdf(request: ExportRequest):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        raise HTTPException(status_code=500, detail="PDF export not available")
    
    businesses = request.businesses
    city = request.city or "Unknown"
    niche = request.niche or "Business"
    
    if not businesses:
        raise HTTPException(status_code=400, detail="No businesses to export")
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, topMargin=50, bottomMargin=50)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#FF6B2B'),
        spaceAfter=10
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.gray,
        spaceAfter=20
    )
    
    elements = []
    
    # Title
    elements.append(Paragraph("LeadRadar", title_style))
    elements.append(Paragraph("Business Lead Finder", subtitle_style))
    elements.append(Paragraph(f"{niche} in {city} - {len(businesses)} Leads", styles['Normal']))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['Name', 'Phone', 'Email', 'Rating']]
    for b in businesses[:50]:  # Limit to 50 for PDF
        data.append([
            b.get('name', '')[:30],
            b.get('phone', ''),
            (b.get('email') or 'N/A')[:25],
            f"{b.get('rating', '')} ({b.get('totalReviews', '')})"
        ])
    
    table = Table(data, colWidths=[150, 100, 150, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6B2B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')])
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Generated by LeadRadar - Premium Business Lead Finder", 
                             ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray)))
    
    doc.build(elements)
    output.seek(0)
    
    return Response(
        content=output.read(),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=leads.pdf"}
    )

# Auth routes - REMINDER: DO NOT HARDCODE THE URL, OR ADD ANY FALLBACKS OR REDIRECT URLS, THIS BREAKS THE AUTH
@app.post("/api/auth/session")
async def create_session(request: SessionRequest):
    session_id = request.sessionId
    
    if not session_id:
        raise HTTPException(status_code=400, detail="Session ID is required")
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(
                "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
                headers={"X-Session-ID": session_id}
            )
            
            if response.status_code != 200:
                raise HTTPException(status_code=401, detail="Authentication failed")
            
            user_data = response.json()
    except Exception as e:
        raise HTTPException(status_code=401, detail="Authentication failed")
    
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    session_token = f"session_{uuid.uuid4()}"
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    
    # Check if user exists
    existing_user = await db.users.find_one(
        {"email": user_data["email"]},
        {"_id": 0}
    )
    
    if not existing_user:
        await db.users.insert_one({
            "user_id": user_id,
            "email": user_data["email"],
            "name": user_data["name"],
            "picture": user_data.get("picture"),
            "plan": "free",
            "leadsToday": 0,
            "lastLeadDate": None,
            "created_at": datetime.now(timezone.utc)
        })
    
    final_user_id = existing_user["user_id"] if existing_user else user_id
    
    # Create session
    await db.user_sessions.insert_one({
        "user_id": final_user_id,
        "session_token": session_token,
        "expires_at": expires_at,
        "created_at": datetime.now(timezone.utc)
    })
    
    user = existing_user or {
        "user_id": user_id,
        "email": user_data["email"],
        "name": user_data["name"],
        "picture": user_data.get("picture"),
        "plan": "free"
    }
    
    response = Response(content='{"user": ' + str(user).replace("'", '"').replace("None", "null") + '}')
    response.headers["Content-Type"] = "application/json"
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7 * 24 * 60 * 60
    )
    
    return response

@app.get("/api/auth/me")
async def get_current_user(request: Request):
    session_token = request.cookies.get("session_token")
    
    if not session_token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            session_token = auth_header[7:]
    
    if not session_token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    session = await db.user_sessions.find_one(
        {"session_token": session_token},
        {"_id": 0}
    )
    
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    # Check expiry
    expires_at = session["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at.replace("Z", "+00:00"))
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expired")
    
    user = await db.users.find_one(
        {"user_id": session["user_id"]},
        {"_id": 0}
    )
    
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    
    return user

@app.post("/api/auth/logout")
async def logout(request: Request, response: Response):
    session_token = request.cookies.get("session_token")
    
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    
    response = Response(content='{"success": true}')
    response.headers["Content-Type"] = "application/json"
    response.delete_cookie(
        key="session_token",
        path="/",
        secure=True,
        samesite="none"
    )
    
    return response
